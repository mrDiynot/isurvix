# test_equipment_insertion.py
from openpyxl import load_workbook
from copy import copy
from openpyxl.utils import get_column_letter

def write_to_cell(ws, cell_ref, value):
    """Write to cell, handling merged cells"""
    from openpyxl.cell.cell import MergedCell
    cell = ws[cell_ref]
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                top_left = merged_range.start_cell
                ws[top_left.coordinate] = value
                return
    else:
        ws[cell_ref] = value

def copy_row_format(ws, source_row, target_row):
    """Copy formatting and merged cells from source to target"""
    # Copy cell styles
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)
        
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
    
    # Copy merged cells
    merged_cells_to_copy = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row == source_row and merged_range.max_row == source_row:
            merged_cells_to_copy.append((merged_range.min_col, merged_range.max_col))
    
    for min_col, max_col in merged_cells_to_copy:
        start_cell = f"{get_column_letter(min_col)}{target_row}"
        end_cell = f"{get_column_letter(max_col)}{target_row}"
        ws.merge_cells(f"{start_cell}:{end_cell}")

# Sample equipment data (what comes from your form)
equipment_data = {
    'STC': {
        'ANTENNA': [
            {'model': 'Antenna Model 1', 'dimension': '1.2m', 'height': '25m', 'azimuth': '90', 'sector': 'A'},
            {'model': 'Antenna Model 2', 'dimension': '1.5m', 'height': '25m', 'azimuth': '180', 'sector': 'B'},
            {'model': 'Antenna Model 3', 'dimension': '1.8m', 'height': '25m', 'azimuth': '270', 'sector': 'C'}
        ],
        'RADIO': [
            {'model': 'Radio RRU 1', 'dimension': 'Small', 'height': '24m', 'sector': 'A'}
        ]
    },
    'OTHER': {
        'ANTENNA': [
            {'model': 'Other Antenna 1', 'dimension': '2.0m', 'height': '30m', 'azimuth': '45', 'sector': 'X'}
        ]
    }
}

# Load your Excel template
template_path = '/path/to/your/template.xlsx'
workbook = load_workbook(template_path)
worksheet = workbook.active

# PROCESS STC ANTENNA (row 198) - 3 antennas
if 'STC' in equipment_data and 'ANTENNA' in equipment_data['STC']:
    start_row = 198
    equipment_list = equipment_data['STC']['ANTENNA']
    
    print(f"Processing {len(equipment_list)} STC ANTENNAs starting at row {start_row}")
    
    # Step 1: Insert rows for additional equipment (beyond first one)
    for idx in range(1, len(equipment_list)):
        print(f"  Inserting row at position {start_row + 1}")
        worksheet.insert_rows(start_row + 1, amount=1)
        copy_row_format(worksheet, start_row, start_row + 1)
    
    # Step 2: Write all equipment data
    for idx, equip in enumerate(equipment_list):
        row = start_row + idx
        print(f"  Writing equipment {idx+1} to row {row}: {equip['model']}")
        write_to_cell(worksheet, f"A{row}", equip.get('model', ''))
        write_to_cell(worksheet, f"C{row}", equip.get('dimension', ''))
        write_to_cell(worksheet, f"D{row}", equip.get('height', ''))
        write_to_cell(worksheet, f"E{row}", equip.get('azimuth', ''))
        write_to_cell(worksheet, f"F{row}", equip.get('sector', ''))

# PROCESS STC RADIO (row 201) - 1 radio (NO azimuth column)
if 'STC' in equipment_data and 'RADIO' in equipment_data['STC']:
    start_row = 201
    equipment_list = equipment_data['STC']['RADIO']
    
    print(f"\nProcessing {len(equipment_list)} STC RADIOs starting at row {start_row}")
    
    # No insertion needed (only 1 item)
    for idx in range(1, len(equipment_list)):
        worksheet.insert_rows(start_row + 1, amount=1)
        copy_row_format(worksheet, start_row, start_row + 1)
    
    for idx, equip in enumerate(equipment_list):
        row = start_row + idx
        print(f"  Writing radio {idx+1} to row {row}: {equip['model']}")
        write_to_cell(worksheet, f"A{row}", equip.get('model', ''))
        write_to_cell(worksheet, f"C{row}", equip.get('dimension', ''))
        write_to_cell(worksheet, f"D{row}", equip.get('height', ''))
        write_to_cell(worksheet, f"E{row}", equip.get('sector', ''))  # Sector in column E (NO azimuth!)

# Save output
workbook.save('/path/to/output.xlsx')
print("\nDone! Excel file saved.")