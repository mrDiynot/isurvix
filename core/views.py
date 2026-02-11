import os
import zipfile
from io import BytesIO

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.core.files.storage import default_storage
from django.db import models, IntegrityError
from django.http import FileResponse, Http404, JsonResponse
from django.shortcuts import get_object_or_404
from django.shortcuts import redirect, render
from django.utils import timezone
from django.utils.text import slugify
from django.views.decorators.http import require_http_methods
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from PIL import Image as PilImage

from .models import Checklist, Profile, Project, GeoLocation, WorkAssignment


def home_view(request):
	"""Homepage view - redirect authenticated users to their dashboard"""
	# If user is logged in, redirect them to their dashboard
	if request.session.get("is_dev_admin"):
		return redirect("dev_admin")
	
	if request.user.is_authenticated:
		profile = getattr(request.user, "profile", None)
		if profile and profile.role == Profile.Roles.ADMIN:
			request.session["is_dev_admin"] = True
			return redirect("dev_admin")
		if profile and profile.path:
			return redirect(f"/{profile.path}")
		if request.user.is_staff or request.user.is_superuser:
			request.session["is_dev_admin"] = True
			return redirect("dev_admin")
	
	# Show public landing page for non-authenticated users
	return render(request, "dashboards/home.html")


@require_http_methods(["GET", "POST"])
def login_view(request):
	if request.session.get("is_dev_admin"):
		return redirect("dev_admin")
	if request.user.is_authenticated:
		profile = getattr(request.user, "profile", None)
		if profile and profile.role == Profile.Roles.ADMIN:
			request.session["is_dev_admin"] = True
			return redirect("dev_admin")
		if profile and profile.path:
			return redirect(f"/{profile.path}")
		# Avoid logging out admin/staff users without a path
		if request.user.is_staff or request.user.is_superuser:
			request.session["is_dev_admin"] = True
			return redirect("dev_admin")
		logout(request)
		return redirect("login")

	if request.method == "POST":
		role = request.POST.get("role")
		username = request.POST.get("username")
		password = request.POST.get("password")

		user_record = User.objects.filter(username=username).first()
		profile_record = getattr(user_record, "profile", None) if user_record else None
		if profile_record and profile_record.is_locked:
			messages.error(request, "Account locked. Contact admin to unlock.")
			return render(request, "login.html", {"selected_role": role, "username": username})

		if role == "ADMIN":
			user = authenticate(request, username=username, password=password)
			if not user:
				if profile_record:
					profile_record.failed_attempts += 1
					remaining = 3 - profile_record.failed_attempts
					if profile_record.failed_attempts >= 3:
						profile_record.is_locked = True
						messages.error(request, "Account locked after 3 failed attempts. Contact admin to unlock.")
					else:
						messages.error(request, f"Incorrect password. {remaining} attempt(s) remaining.")
					profile_record.save(update_fields=["failed_attempts", "is_locked"])
				else:
					messages.error(request, "Username not found.")
				return render(request, "login.html", {"selected_role": role, "username": username})
			profile = getattr(user, "profile", None)
			if not profile or profile.role != Profile.Roles.ADMIN:
				messages.error(request, "This account is not an App Admin.")
				return render(request, "login.html", {"selected_role": role, "username": username})
			# Ensure admin users can access Django admin panel
			updated_fields = []
			if not user.is_staff:
				user.is_staff = True
				updated_fields.append("is_staff")
			if not user.is_superuser:
				user.is_superuser = True
				updated_fields.append("is_superuser")
			if updated_fields:
				user.save(update_fields=updated_fields)
			profile.failed_attempts = 0
			profile.is_locked = False
			profile.save(update_fields=["failed_attempts", "is_locked"])
			login(request, user)
			request.session["is_dev_admin"] = True
			return redirect("dev_admin")

		user = authenticate(request, username=username, password=password)
		if not user:
			if profile_record:
				profile_record.failed_attempts += 1
				remaining = 3 - profile_record.failed_attempts
				if profile_record.failed_attempts >= 3:
					profile_record.is_locked = True
					messages.error(request, "Account locked after 3 failed attempts. Contact admin to unlock.")
				else:
					messages.error(request, f"Incorrect password. {remaining} attempt(s) remaining.")
				profile_record.save(update_fields=["failed_attempts", "is_locked"])
			else:
				messages.error(request, "Username not found.")
			return render(request, "login.html", {"selected_role": role, "username": username})

		try:
			profile = user.profile
		except Profile.DoesNotExist:
			messages.error(request, "No profile found for this user.")
			return render(request, "login.html", {"selected_role": role, "username": username})

		if role == "TEAM_LEAD" and profile.role != Profile.Roles.TEAM_LEAD:
			messages.error(request, "This account is not a Team Leader.")
			return render(request, "login.html", {"selected_role": role, "username": username})

		if role == "ENGINEER" and profile.role != Profile.Roles.ENGINEER:
			messages.error(request, "This account is not an Engineer.")
			return render(request, "login.html", {"selected_role": role, "username": username})

		if not profile.path:
			messages.error(request, "Your account is missing a path. Ask admin to set it.")
			return render(request, "login.html", {"selected_role": role, "username": username})

		login(request, user)
		profile.failed_attempts = 0
		profile.is_locked = False
		profile.save(update_fields=["failed_attempts", "is_locked"])
		return redirect(f"/{profile.path}")

	return render(request, "login.html")


@require_http_methods(["POST"])
def logout_view(request):
	request.session.pop("is_dev_admin", None)
	logout(request)
	return redirect("login")


@require_http_methods(["GET", "POST"])
def dev_admin_view(request):
	if not request.session.get("is_dev_admin"):
		profile = getattr(request.user, "profile", None) if request.user.is_authenticated else None
		if profile and profile.role == Profile.Roles.ADMIN:
			request.session["is_dev_admin"] = True
		elif request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser):
			request.session["is_dev_admin"] = True
		else:
			messages.error(request, "Please log in as admin.")
			return redirect("login")

	if request.method == "POST":
		action = request.POST.get("action")
		if action == "create_project":
			name = request.POST.get("project_name", "").strip()
			description = request.POST.get("project_description", "").strip()
			if not name:
				messages.error(request, "Project name is required.")
				return redirect("dev_admin")
			if Project.objects.filter(name__iexact=name).exists():
				messages.error(request, "Project name already exists.")
				return redirect("dev_admin")
			Project.objects.create(name=name, description=description)
			messages.success(request, f"Project {name} created.")
			return redirect("dev_admin")

		if action == "create_user":
			username = request.POST.get("username", "").strip()
			password = request.POST.get("password", "").strip()
			role = request.POST.get("role")
			path = request.POST.get("path", "").strip().lstrip("/")
			project_id = request.POST.get("project")

			if not username or not password or not role or not path or not project_id:
				messages.error(request, "All fields are required.")
				return redirect("dev_admin")

			if role not in (Profile.Roles.TEAM_LEAD, Profile.Roles.ENGINEER):
				messages.error(request, "Role must be Team Leader or Engineer.")
				return redirect("dev_admin")

			# Case-insensitive path validation
			path_lower = path.lower()
			if role == Profile.Roles.TEAM_LEAD and not path_lower.startswith("tl"):
				messages.error(request, "Team Leader path must start with TL (e.g. TL1).")
				return redirect("dev_admin")

			if role == Profile.Roles.ENGINEER and not path_lower.startswith("eng"):
				messages.error(request, "Engineer path must start with Eng (e.g. Eng1).")
				return redirect("dev_admin")

			if User.objects.filter(username=username).exists():
				messages.error(request, "Username already exists.")
				return redirect("dev_admin")

			if Profile.objects.filter(path=path).exists():
				messages.error(request, "Path already in use.")
				return redirect("dev_admin")

			project = Project.objects.filter(id=project_id).first()
			if not project:
				messages.error(request, "Selected project not found.")
				return redirect("dev_admin")

			user = User.objects.create_user(username=username, password=password)
			profile = user.profile
			profile.role = role
			profile.path = path
			profile.project = project
			profile.save()

			messages.success(request, f"User {username} created with path /{path}.")
			return redirect("dev_admin")

		messages.error(request, "Invalid action.")
		return redirect("dev_admin")

	# Get filter parameters
	status_filter = request.GET.get("status", "")
	user_filter = request.GET.get("user", "")
	search_query = request.GET.get("q", "").strip()

	users = User.objects.select_related("profile", "profile__project").filter(is_superuser=False).order_by("username")
	projects = Project.objects.order_by("name")
	
	checklists = Checklist.objects.select_related("user", "user__profile", "project")
	
	# Apply filters
	if status_filter:
		checklists = checklists.filter(status=status_filter)
	if user_filter:
		checklists = checklists.filter(user_id=user_filter)
	if search_query:
		checklists = checklists.filter(site_id__icontains=search_query)
	
	checklists = checklists.order_by("-updated_at")
	
	# Get all engineers for filter dropdown
	engineers = User.objects.filter(
		profile__role=Profile.Roles.ENGINEER
	).order_by("username")
	
	locked_profiles = Profile.objects.select_related("user", "project").filter(is_locked=True).order_by("user__username")
	user_stats = (
		Checklist.objects.values("user__username")
		.filter(status=Checklist.Status.FINAL)
		.annotate(total=models.Count("id"))
		.order_by("-total")
	)
	
	# Get all locations that are NOT already assigned
	assigned_site_ids = WorkAssignment.objects.values_list('site_id', flat=True)
	locations = GeoLocation.objects.select_related("project", "created_by").exclude(
		name__in=assigned_site_ids
	).order_by("-created_at")
	
	# Get all work assignments
	work_assignments = WorkAssignment.objects.select_related(
		"assigned_to", "assigned_to__profile", "assigned_by", "project"
	).order_by("-created_at")
	
	return render(
		request,
		"dashboards/dev_admin.html",
		{
			"users": users,
			"projects": projects,
			"checklists": checklists,
			"locked_profiles": locked_profiles,
			"user_stats": user_stats,
			"engineers": engineers,
			"status_filter": status_filter,
			"user_filter": user_filter,
			"search_query": search_query,
			"locations": locations,
			"work_assignments": work_assignments,
		},
	)


@require_http_methods(["GET", "POST"])
def admin_user_edit(request, user_id: int):
	if not request.session.get("is_dev_admin"):
		messages.error(request, "Please log in as admin.")
		return redirect("login")

	user = get_object_or_404(User, pk=user_id)
	profile = getattr(user, "profile", None)
	if not profile:
		messages.error(request, "Profile not found.")
		return redirect("dev_admin")

	if request.method == "POST":
		username = request.POST.get("username", "").strip()
		password = request.POST.get("password", "").strip()
		role = request.POST.get("role")
		path = request.POST.get("path", "").strip().lstrip("/")
		project_id = request.POST.get("project")

		if not username or not role or not path or not project_id:
			messages.error(request, "Username, role, path, and project are required.")
			return redirect("admin_user_edit", user_id=user.id)

		if User.objects.filter(username=username).exclude(id=user.id).exists():
			messages.error(request, "Username already exists.")
			return redirect("admin_user_edit", user_id=user.id)

		if Profile.objects.filter(path=path).exclude(user=user).exists():
			messages.error(request, "Path already in use.")
			return redirect("admin_user_edit", user_id=user.id)

		# Case-insensitive path validation
		path_lower = path.lower()
		if role == Profile.Roles.TEAM_LEAD and not path_lower.startswith("tl"):
			messages.error(request, "Team Leader path must start with TL (e.g. TL1).")
			return redirect("admin_user_edit", user_id=user.id)

		if role == Profile.Roles.ENGINEER and not path_lower.startswith("eng"):
			messages.error(request, "Engineer path must start with Eng (e.g. Eng1).")
			return redirect("admin_user_edit", user_id=user.id)

		project = Project.objects.filter(id=project_id).first()
		if not project:
			messages.error(request, "Selected project not found.")
			return redirect("admin_user_edit", user_id=user.id)

		user.username = username
		if password:
			user.set_password(password)
		user.save()

		profile.role = role
		profile.path = path
		profile.project = project
		profile.save()

		messages.success(request, "User updated.")
		return redirect("dev_admin")

	projects = Project.objects.order_by("name")
	return render(
		request,
		"dashboards/admin_user_edit.html",
		{"edit_user": user, "profile": profile, "projects": projects},
	)


@require_http_methods(["POST"])
def admin_user_delete(request, user_id: int):
	if not request.session.get("is_dev_admin"):
		messages.error(request, "Please log in as admin.")
		return redirect("login")

	user = get_object_or_404(User, pk=user_id)
	if user.is_superuser:
		messages.error(request, "Cannot delete a superuser.")
		return redirect("dev_admin")

	user.delete()
	messages.success(request, "User deleted.")
	return redirect("dev_admin")


@require_http_methods(["POST"])
def admin_user_unlock(request, profile_id: int):
	if not request.session.get("is_dev_admin"):
		messages.error(request, "Please log in as admin.")
		return redirect("login")

	profile = get_object_or_404(Profile, id=profile_id)
	profile.is_locked = False
	profile.failed_attempts = 0
	profile.save(update_fields=["is_locked", "failed_attempts"])
	messages.success(request, f"{profile.user.username} unlocked.")
	return redirect("dev_admin")


def user_dashboard(request, path: str):
	"""Unified dashboard that routes to team lead or engineer view based on user role"""
	if not request.user.is_authenticated:
		messages.error(request, "Please log in.")
		return redirect("login")

	profile = getattr(request.user, "profile", None)
	if not profile or profile.path != path:
		messages.error(request, "Access denied.")
		return redirect("login")

	if profile.role == Profile.Roles.TEAM_LEAD:
		return _team_lead_dashboard(request, profile, path)
	elif profile.role == Profile.Roles.ENGINEER:
		return _engineer_dashboard(request, profile, path)
	else:
		messages.error(request, "Invalid user role.")
		return redirect("login")


def _team_lead_dashboard(request, profile, path):
	"""Team lead dashboard logic"""
	# Get filter parameters
	status_filter = request.GET.get("status", "")
	user_filter = request.GET.get("user", "")
	search_query = request.GET.get("q", "").strip()
	
	checklists = Checklist.objects.select_related("user", "user__profile", "project").filter(project=profile.project)
	
	# Apply filters
	if status_filter:
		checklists = checklists.filter(status=status_filter)
	if user_filter:
		checklists = checklists.filter(user_id=user_filter)
	if search_query:
		checklists = checklists.filter(site_id__icontains=search_query)
	
	checklists = checklists.order_by("-updated_at")
	
	# Get users in this project for filter dropdown
	project_users = User.objects.filter(
		profile__project=profile.project,
		profile__role=Profile.Roles.ENGINEER
	).order_by("username")
	
	user_stats = (
		Checklist.objects.values("user__username")
		.filter(project=profile.project, status=Checklist.Status.FINAL)
		.annotate(total=models.Count("id"))
		.order_by("-total")
	)
	
	# Get locations for team lead's project that are NOT already assigned
	assigned_site_ids = WorkAssignment.objects.values_list('site_id', flat=True)
	locations = GeoLocation.objects.select_related("project", "created_by").filter(
		models.Q(project=profile.project) | models.Q(project__isnull=True)
	).exclude(
		name__in=assigned_site_ids
	).order_by("-created_at")
	
	# Get work assignments for team lead's project
	work_assignments = WorkAssignment.objects.select_related(
		"assigned_to", "assigned_to__profile", "assigned_by", "project"
	).filter(project=profile.project).order_by("-created_at")
	
	# Get engineers in this project for assignment form
	engineers = User.objects.filter(
		profile__project=profile.project,
		profile__role=Profile.Roles.ENGINEER
	).order_by("username")

	return render(
		request,
		"dashboards/team_lead.html",
		{
			"path": path,
			"project": profile.project,
			"checklists": checklists,
			"user_stats": user_stats,
			"project_users": project_users,
			"status_filter": status_filter,
			"user_filter": user_filter,
			"search_query": search_query,
			"locations": locations,
			"work_assignments": work_assignments,
			"engineers": engineers,
		},
	)


def _engineer_dashboard(request, profile, path):
	"""Engineer dashboard logic"""
	query = request.GET.get("q", "").strip()
	checklists = (
		Checklist.objects.filter(user=request.user, project=profile.project)
		.order_by("-updated_at")
	)
	if query:
		checklists = checklists.filter(site_id__icontains=query)
	
	# Get work assignments for this engineer
	my_work = WorkAssignment.objects.select_related(
		"assigned_by", "project"
	).filter(assigned_to=request.user).order_by("-created_at")

	# Auto-create draft checklist for any assigned work without one
	for work in my_work:
		if not work.checklist:
			answer_data = {"12": work.site_id}
			checklist = Checklist.objects.create(
				user=request.user,
				project=work.project,
				site_id=work.site_id,
				status=Checklist.Status.DRAFT,
				answer_data=answer_data,
			)
			work.checklist = checklist
			work.save(update_fields=["checklist"])
			_create_or_update_excel_copy(checklist)

	return render(
		request,
		"dashboards/engineer.html",
		{
			"path": path,
			"idx": 0,  # Keep for template compatibility
			"project": profile.project,
			"checklists": checklists,
			"query": query,
			"my_work": my_work,
		},
	)


def team_lead_view(request, idx: int):
	if not request.user.is_authenticated:
		messages.error(request, "Please log in.")
		return redirect("login")

	profile = getattr(request.user, "profile", None)
	expected_path = f"TL{idx}"
	if not profile or profile.role != Profile.Roles.TEAM_LEAD or profile.path != expected_path:
		messages.error(request, "Access denied.")
		return redirect("login")

	checklists = Checklist.objects.select_related("user", "project").filter(project=profile.project).order_by("-updated_at")
	user_stats = (
		Checklist.objects.values("user__username")
		.filter(project=profile.project, status=Checklist.Status.FINAL)
		.annotate(total=models.Count("id"))
		.order_by("-total")
	)

	projects = Project.objects.order_by("name")
	return render(
		request,
		"dashboards/team_lead.html",
		{
			"path": expected_path,
			"project": profile.project,
			"checklists": checklists,
			"user_stats": user_stats,
			"projects": projects,
		},
	)


def engineer_view(request, idx: int):
	profile, redirect_response = _ensure_engineer_access(request, idx)
	if redirect_response:
		return redirect_response

	query = request.GET.get("q", "").strip()
	checklists = (
		Checklist.objects.filter(user=request.user, project=profile.project)
		.order_by("-updated_at")
	)
	if query:
		checklists = checklists.filter(site_id__icontains=query)

	return render(
		request,
		"dashboards/engineer.html",
		{
			"path": profile.path,
			"idx": idx,
			"project": profile.project,
			"checklists": checklists,
			"query": query,
		},
	)


def engineer_checklist_new(request, path: str):
	profile, redirect_response = _ensure_engineer_access(request, path)
	if redirect_response:
		return redirect_response

	project = profile.project
	if not project or not project.template_file:
		messages.error(request, "Your project does not have a template file yet.")
		return redirect("user_dashboard", path=path)

	general_questions, image_questions, dc_power_questions, _site_id = _read_template_questions(project.template_file.path)
	checklist = Checklist.objects.create(user=request.user, project=project)
	_create_or_update_excel_copy(checklist)

	return redirect("checklist_detail", checklist_id=checklist.id)


@require_http_methods(["GET"])
def engineer_checklist_edit(request, path: str, checklist_id: int):
	access = _get_checklist_access(request, checklist_id, path)
	if access["redirect"]:
		return access["redirect"]
	profile = access.get("profile")
	checklist = access["checklist"]
	project = profile.project
	
	# Use checklist's own template copy if it exists, otherwise use project template
	if checklist.template_copy:
		template_path = checklist.template_copy.path
	elif project.template_file:
		template_path = project.template_file.path
	else:
		messages.error(request, "Your project does not have a template file yet.")
		return redirect("user_dashboard", path=path)

	general_questions, image_questions, dc_power_questions, _site_id = _read_template_questions(template_path)

	answers = checklist.answer_data or {}
	remarks = checklist.remark_data or {}
	images = checklist.image_data or {}

	questions = [
		{"row": q["row"], "text": q["text"], "value": answers.get(str(q["row"]), "")}
		for q in general_questions
	]
	questions.extend([
		{"row": q["row"], "text": q["text"], "value": answers.get(str(q["row"]), "")}
		for q in dc_power_questions
	])
	image_questions = [
		{
			"row": q["row"],
			"text": q["text"],
			"remark": remarks.get(str(q["row"]), ""),
			"images": images.get(str(q["row"]), []),
		}
		for q in image_questions
	]

	# Redirect to checklist_detail_view instead of using engineer_checklist_edit template
	return redirect("checklist_detail", checklist_id=checklist_id)


@require_http_methods(["POST"])
def engineer_checklist_autosave(request, path: str, checklist_id: int):
	access = _get_checklist_access(request, checklist_id, path)
	if access["redirect"]:
		return access["redirect"]
	checklist = access["checklist"]
	if not access["can_edit"]:
		return JsonResponse({"status": "locked"}, status=403)

	answers = checklist.answer_data or {}
	remarks = checklist.remark_data or {}
	images = checklist.image_data or {}

	for row in range(4, 19):
		answers[str(row)] = request.POST.get(f"answer_{row}", "").strip()

	site_id_value = answers.get("12", "").strip()
	if site_id_value:
		checklist.site_id = site_id_value

	for row in range(21, 23):
		remarks[str(row)] = request.POST.get(f"remark_{row}", "").strip()
		upload_key = f"images_{row}"
		if upload_key in request.FILES:
			uploads = request.FILES.getlist(upload_key)
			row_images = images.get(str(row), [])
			for upload in uploads:
				file_name = default_storage.save(_build_image_path(checklist, upload.name), upload)
				row_images.append(file_name)
			images[str(row)] = row_images

	checklist.answer_data = answers
	checklist.remark_data = remarks
	checklist.image_data = images
	update_fields = ["answer_data", "remark_data", "image_data", "updated_at"]
	if site_id_value:
		update_fields.append("site_id")
	checklist.save(update_fields=update_fields)
	_create_or_update_excel_copy(checklist)

	return JsonResponse({"status": "ok", "updated": True})


@require_http_methods(["POST"])
def engineer_checklist_delete(request, path: str, checklist_id: int):
	access = _get_checklist_access(request, checklist_id, path=path)
	if access["redirect"]:
		return access["redirect"]
	checklist = access["checklist"]
	if access.get("role") == Profile.Roles.ENGINEER:
		messages.error(request, "Engineers cannot delete checklists.")
		return redirect("engineer_checklist_edit", path=path, checklist_id=checklist.id)

	if checklist.template_copy and default_storage.exists(checklist.template_copy.name):
		default_storage.delete(checklist.template_copy.name)

	for image_list in (checklist.image_data or {}).values():
		for image_path in image_list:
			if default_storage.exists(image_path):
				default_storage.delete(image_path)

	checklist.delete()
	messages.success(request, "Checklist deleted.")
	if access.get("back_path"):
		return redirect(access["back_url"], path=access["back_path"])
	return redirect(access["back_url"])


@require_http_methods(["POST"])
def engineer_checklist_submit(request, path: str, checklist_id: int):
	from django.utils import timezone
	
	access = _get_checklist_access(request, checklist_id, path=path)
	if access["redirect"]:
		return access["redirect"]
	checklist = access["checklist"]
	if access.get("role") != Profile.Roles.ENGINEER:
		messages.error(request, "Only engineers can submit.")
		return redirect("engineer_checklist_edit", path=path, checklist_id=checklist.id)

	checklist.status = Checklist.Status.SUBMITTED
	checklist.save(update_fields=["status", "updated_at"])
	
	# Update linked work assignment status to SUBMITTED
	try:
		work_assignment = WorkAssignment.objects.get(checklist=checklist)
		work_assignment.status = WorkAssignment.Status.SUBMITTED
		work_assignment.submitted_at = timezone.now()
		work_assignment.save()
	except WorkAssignment.DoesNotExist:
		pass  # No linked work assignment
	
	messages.success(request, "Checklist submitted for review.")
	return redirect("engineer_checklist_edit", path=path, checklist_id=checklist.id)


@require_http_methods(["POST"])
def checklist_review_update(request, checklist_id: int):
	access = _get_checklist_access(request, checklist_id)
	if access["redirect"]:
		return access["redirect"]
	if access.get("role") not in ("ADMIN", Profile.Roles.TEAM_LEAD):
		messages.error(request, "Only admin or team lead can update status.")
		return redirect("login")

	checklist = access["checklist"]
	status = request.POST.get("status")
	comment = request.POST.get("comment", "").strip()

	if status and status in [
		Checklist.Status.REVIEW,
		Checklist.Status.FINAL,
		Checklist.Status.SUBMITTED,
		Checklist.Status.DRAFT,
	]:
		checklist.status = status
		
		# Update work assignment status if checklist is marked as FINAL
		if status == Checklist.Status.FINAL:
			try:
				work_assignment = WorkAssignment.objects.get(checklist=checklist)
				work_assignment.status = WorkAssignment.Status.COMPLETED
				work_assignment.completed_at = timezone.now()
				work_assignment.save(update_fields=["status", "completed_at"])
			except WorkAssignment.DoesNotExist:
				pass

	# Save comment and track who added it
	if comment != checklist.comment:
		checklist.comment = comment
		checklist.comment_by = request.user
		checklist.save(update_fields=["status", "comment", "comment_by", "updated_at"])
	else:
		checklist.save(update_fields=["status", "updated_at"])
	messages.success(request, "Checklist updated.")
	if access.get("back_path"):
		return redirect(access["back_url"], path=access["back_path"])
	return redirect(access["back_url"])


def engineer_checklist_download(request, path: str, checklist_id: int):
	access = _get_checklist_access(request, checklist_id, path)
	if access["redirect"]:
		return access["redirect"]
	checklist = access["checklist"]
	if not checklist.template_copy:
		_create_or_update_excel_copy(checklist)

	if not checklist.template_copy or not default_storage.exists(checklist.template_copy.name):
		raise Http404("Checklist file not found.")

	file_handle = default_storage.open(checklist.template_copy.name, "rb")
	# Use site_id in download filename
	site_id_slug = _safe_slug(checklist.site_id, f"site_{checklist.id}")
	filename = f"{site_id_slug}.xlsx"
	response = FileResponse(file_handle, as_attachment=True, filename=filename)
	return response


def _ensure_engineer_access(request, path: str):
	if not request.user.is_authenticated:
		messages.error(request, "Please log in.")
		return None, redirect("login")

	profile = getattr(request.user, "profile", None)
	if not profile or profile.role != Profile.Roles.ENGINEER or profile.path != path:
		messages.error(request, "Access denied.")
		return None, redirect("login")

	return profile, None


def _get_checklist_access(request, checklist_id: int, path: str | None = None):
	if request.session.get("is_dev_admin"):
		checklist = get_object_or_404(Checklist, id=checklist_id)
		return {
			"checklist": checklist,
			"profile": Profile(user=checklist.user, role=Profile.Roles.ENGINEER, project=checklist.project, path=""),
			"can_edit": True,
			"role": "ADMIN",
			"redirect": None,
			"back_url": "dev_admin",
			"back_name": "dev_admin",
			"back_path": None,
		}

	if not request.user.is_authenticated:
		messages.error(request, "Please log in.")
		return {"redirect": redirect("login")}

	profile = getattr(request.user, "profile", None)
	if not profile:
		messages.error(request, "Profile not found.")
		return {"redirect": redirect("login")}

	if profile.role == Profile.Roles.ENGINEER:
		checklist = get_object_or_404(Checklist, id=checklist_id, user=request.user, project=profile.project)
		if path is not None and profile.path != path:
			messages.error(request, "Access denied.")
			return {"redirect": redirect("login")}
		return {
			"checklist": checklist,
			"profile": profile,
			"can_edit": checklist.status != Checklist.Status.FINAL,
			"role": profile.role,
			"redirect": None,
			"back_url": "user_dashboard",
			"back_name": profile.path,
			"back_path": profile.path,
		}

	if profile.role == Profile.Roles.TEAM_LEAD:
		checklist = get_object_or_404(Checklist, id=checklist_id, project=profile.project)
		return {
			"checklist": checklist,
			"profile": profile,
			"can_edit": True,
			"role": profile.role,
			"redirect": None,
			"back_url": "user_dashboard",
			"back_name": profile.path,
			"back_path": profile.path,
		}

	messages.error(request, "Access denied.")
	return {"redirect": redirect("login")}


def _read_template_questions(template_path: str):
	workbook = load_workbook(template_path)
	worksheet = workbook.active

	# General section: rows 4-18, Questions in AB (merged), Answers in CDEF (merged)
	general_questions = []
	for row in range(4, 19):
		text = worksheet[f"A{row}"].value or ""
		if text and str(text).strip():
			general_questions.append({"row": row, "text": str(text).strip()})

	# All photo sections: Questions in B, Remarks in DE (merged), Images in F, G, H...
	image_questions = []
	# CIVIL & SITE GENERAL (rows 22-70)
	for row in range(22, 71):
		text = worksheet[f"B{row}"].value or ""
		if text and str(text).strip():
			image_questions.append({"row": row, "text": str(text).strip()})
	
	# ELECTROMECHANICAL (rows 73-84)
	for row in range(73, 85):
		text = worksheet[f"B{row}"].value or ""
		if text and str(text).strip():
			image_questions.append({"row": row, "text": str(text).strip()})
	
	# Power Supply (rows 87-96)
	for row in range(87, 97):
		text = worksheet[f"B{row}"].value or ""
		if text and str(text).strip():
			image_questions.append({"row": row, "text": str(text).strip()})
	
	# DG/SG Set (rows 98-114)
	for row in range(98, 115):
		text = worksheet[f"B{row}"].value or ""
		if text and str(text).strip():
			image_questions.append({"row": row, "text": str(text).strip()})
	
	# Hybrid AC/DC (rows 116-125)
	for row in range(116, 126):
		text = worksheet[f"B{row}"].value or ""
		if text and str(text).strip():
			image_questions.append({"row": row, "text": str(text).strip()})
	
	# Solar (rows 127-133)
	for row in range(127, 134):
		text = worksheet[f"B{row}"].value or ""
		if text and str(text).strip():
			image_questions.append({"row": row, "text": str(text).strip()})
	
	# Transformer (rows 135-142)
	for row in range(135, 143):
		text = worksheet[f"B{row}"].value or ""
		if text and str(text).strip():
			image_questions.append({"row": row, "text": str(text).strip()})
	
	# Service Disconnect (rows 144-151)
	for row in range(144, 152):
		text = worksheet[f"B{row}"].value or ""
		if text and str(text).strip():
			image_questions.append({"row": row, "text": str(text).strip()})
	
	# SHAREABLE MDB/LDB (rows 153-160)
	for row in range(153, 161):
		text = worksheet[f"B{row}"].value or ""
		if text and str(text).strip():
			image_questions.append({"row": row, "text": str(text).strip()})
	
	# SHELTER/ODU (rows 163-181)
	for row in range(163, 182):
		text = worksheet[f"B{row}"].value or ""
		if text and str(text).strip():
			image_questions.append({"row": row, "text": str(text).strip()})
	
	# Extra (rows 184-185)
	for row in range(184, 186):
		text = worksheet[f"B{row}"].value or ""
		if text and str(text).strip():
			image_questions.append({"row": row, "text": str(text).strip()})

	# DC Power System Data (rows 187-193): Labels in AB (merged), Values in DEF (merged)
	dc_power_questions = []
	for row in range(187, 194):
		text = worksheet[f"A{row}"].value or ""
		if text and str(text).strip():
			dc_power_questions.append({"row": row, "text": str(text).strip()})

	site_id = worksheet["B12"].value or worksheet["A12"].value or ""
	return general_questions, image_questions, dc_power_questions, str(site_id).strip()


def _create_or_update_excel_copy(checklist: Checklist):
	"""
	Create or update Excel copy for checklist.
	ALWAYS uses the current project template for new checklists.
	"""
	project = checklist.project
	if not project or not project.template_file:
		return

	# ALWAYS start from the project template to avoid duplicate images
	# This ensures we have a clean slate each time
	template_path = project.template_file.path
	
	workbook = load_workbook(template_path)
	worksheet = workbook.active
	# Add template images (logos, headers, etc.) from the original template
	_add_template_images(worksheet, template_path)

	def write_to_cell(ws, cell_ref, value):
		"""Safely write to a cell, handling merged cells"""
		from openpyxl.cell.cell import MergedCell
		cell = ws[cell_ref]
		if isinstance(cell, MergedCell):
			# Find the top-left cell of the merged range
			for merged_range in ws.merged_cells.ranges:
				if cell.coordinate in merged_range:
					top_left = merged_range.start_cell
					ws[top_left.coordinate] = value
					return
		else:
			ws[cell_ref] = value

	def copy_row_format(ws, source_row, target_row):
		"""Copy formatting and merged cells from source row to target row"""
		from copy import copy
		from openpyxl.utils import get_column_letter
		
		# Copy cell formatting
		for col in range(1, ws.max_column + 1):
			source_cell = ws.cell(row=source_row, column=col)
			target_cell = ws.cell(row=target_row, column=col)
			
			if source_cell.has_style:
				target_cell.font = copy(source_cell.font)
				target_cell.border = copy(source_cell.border)
				target_cell.fill = copy(source_cell.fill)
				target_cell.number_format = copy(source_cell.number_format)
				target_cell.protection = copy(source_cell.protection)
				target_cell.alignment = copy(source_cell.alignment)
		
		# Copy merged cells - check if source row has any merged cells
		merged_cells_to_copy = []
		for merged_range in list(ws.merged_cells.ranges):
			if merged_range.min_row == source_row and merged_range.max_row == source_row:
				# This is a merged cell in the source row
				merged_cells_to_copy.append((merged_range.min_col, merged_range.max_col))
		
		# Apply merged cells to target row
		for min_col, max_col in merged_cells_to_copy:
			start_cell = f"{get_column_letter(min_col)}{target_row}"
			end_cell = f"{get_column_letter(max_col)}{target_row}"
			ws.merge_cells(f"{start_cell}:{end_cell}")
	
	def clear_row_style(ws, row):
		"""Remove all formatting from row (fixes openpyxl auto-copy issue)"""
		from openpyxl.styles import PatternFill
		
		for col in range(1, ws.max_column + 1):
			cell = ws.cell(row=row, column=col)
			cell.font = None
			cell.border = None
			cell.fill = PatternFill()  # Remove background color
			cell.number_format = 'General'
			cell.alignment = None
			cell.protection = None
	
	def insert_clean_row(ws, template_row, insert_at):
		"""Insert row like Excel and apply template format"""
		# Step 1: Insert row
		ws.insert_rows(insert_at)
		# Step 2: Remove any merged ranges that shifted onto this row
		template_merges = set()
		for merged_range in list(ws.merged_cells.ranges):
			if merged_range.min_row == template_row and merged_range.max_row == template_row:
				template_merges.add((merged_range.min_col, merged_range.max_col))
			elif merged_range.min_row <= insert_at <= merged_range.max_row:
				try:
					ws.unmerge_cells(str(merged_range))
				except KeyError:
					# Some cells may already be missing in openpyxl's internal map
					pass
		# Step 3: Clear auto-copied formatting
		clear_row_style(ws, insert_at)
		# Step 4: Copy correct format from template
		copy_row_format(ws, template_row, insert_at)

	answers = checklist.answer_data or {}
	# General section: Write answers to CDEF merged cells (rows 4-18)
	for row in range(4, 19):
		value = answers.get(str(row), "")
		if value:
			# CDEF is merged, so write to C (first column of merge)
			write_to_cell(worksheet, f"C{row}", value)

	# DC Power System: Write values to DEF merged cells (rows 187-193)
	for row in range(187, 194):
		value = answers.get(str(row), "")
		if value:
			# DEF is merged, so write to D (first column of merge)
			write_to_cell(worksheet, f"D{row}", value)

	# Write tower equipment data to Excel
	# Extract equipment data from answer_data
	equipment_data = {}
	electrical_data = {}
	
	# Debug: Print what we're extracting
	import sys
	print(f"\n=== DEBUG: Processing checklist {checklist.id} ===", file=sys.stderr)
	print(f"Total answer_data keys: {len(answers)}", file=sys.stderr)
	
	for key, value in answers.items():
		if key.startswith('equipment_'):
			print(f"Found equipment key: {key}", file=sys.stderr)
			print(f"Value: {value}", file=sys.stderr)
			
			operator = value.get('operator', '')
			equip_type = value.get('type', '')
			equip_info = value.get('data', {})
			
			if operator not in equipment_data:
				equipment_data[operator] = {}
			if equip_type not in equipment_data[operator]:
				equipment_data[operator][equip_type] = []
			equipment_data[operator][equip_type].append(equip_info)
			
		elif key.startswith('electrical_'):
			row_num = int(key.replace('electrical_', ''))
			electrical_data[row_num] = value
	
	print(f"Extracted equipment_data: {equipment_data}", file=sys.stderr)
	print(f"Extracted electrical_data: {electrical_data}", file=sys.stderr)
	
	# Write equipment data to Excel (STATIC rows, no insert)
	def _sorted_equipment_list(items):
		return sorted(items, key=lambda x: int(x.get('position_index', 0) or 0))
	
	def _write_equipment_block(operator, equip_type, start_row, max_rows, col_map):
		if operator not in equipment_data:
			return
		if equip_type not in equipment_data[operator]:
			return
		items = _sorted_equipment_list(equipment_data[operator][equip_type])[:max_rows]
		for idx, equip in enumerate(items):
			row = start_row + idx
			if 'model' in col_map:
				write_to_cell(worksheet, f"{col_map['model']}{row}", equip.get('model', ''))
			if 'dimension' in col_map:
				write_to_cell(worksheet, f"{col_map['dimension']}{row}", equip.get('dimension', ''))
			if 'height' in col_map:
				write_to_cell(worksheet, f"{col_map['height']}{row}", equip.get('height', ''))
			if 'azimuth' in col_map:
				write_to_cell(worksheet, f"{col_map['azimuth']}{row}", equip.get('azimuth', ''))
			if 'empty_port' in col_map:
				write_to_cell(worksheet, f"{col_map['empty_port']}{row}", equip.get('empty_port', ''))
			if 'sector' in col_map:
				write_to_cell(worksheet, f"{col_map['sector']}{row}", equip.get('sector', ''))

	# STC columns: AB merged => A, C, D, E, F
	_write_equipment_block('STC', 'ANTENNA', 198, 15, {
		'model': 'A', 'dimension': 'C', 'height': 'D', 'azimuth': 'E', 'sector': 'F'
	})
	_write_equipment_block('STC', 'RADIO', 215, 15, {
		'model': 'A', 'dimension': 'C', 'height': 'D', 'sector': 'E'
	})
	_write_equipment_block('STC', 'FPFH', 232, 15, {
		'model': 'A', 'dimension': 'C', 'height': 'D', 'empty_port': 'E', 'sector': 'F'
	})
	_write_equipment_block('STC', 'MICROWAVE', 249, 9, {
		'model': 'A', 'dimension': 'C', 'height': 'D', 'azimuth': 'E', 'sector': 'F'
	})

	# OTHER operator columns: IJ merged => I, K, L, M, N
	_write_equipment_block('OTHER', 'ANTENNA', 198, 18, {
		'model': 'I', 'dimension': 'K', 'height': 'L', 'azimuth': 'M', 'sector': 'N'
	})
	_write_equipment_block('OTHER', 'RADIO', 218, 18, {
		'model': 'I', 'dimension': 'K', 'height': 'L', 'sector': 'M'
	})
	_write_equipment_block('OTHER', 'FPFH', 238, 18, {
		'model': 'I', 'dimension': 'K', 'height': 'L', 'empty_port': 'M', 'sector': 'N'
	})
	_write_equipment_block('OTHER', 'MICROWAVE', 258, 9, {
		'model': 'I', 'dimension': 'K', 'height': 'L', 'azimuth': 'M', 'sector': 'N'
	})

	# Write electrical data (rows 261-263)
	for row_num, elec in electrical_data.items():
		if 261 <= row_num <= 263:
			write_to_cell(worksheet, f"A{row_num}", elec.get('voltage', ''))
			write_to_cell(worksheet, f"C{row_num}", elec.get('current_r', ''))
			write_to_cell(worksheet, f"D{row_num}", elec.get('current_y', ''))
			write_to_cell(worksheet, f"E{row_num}", elec.get('current_b', ''))
			write_to_cell(worksheet, f"F{row_num}", elec.get('remarks', ''))
	
	remarks = checklist.remark_data or {}
	images = checklist.image_data or {}
	# Photo sections: Write remarks to DE merged cells and images starting from F
	all_photo_rows = (list(range(22, 71)) + list(range(73, 85)) + list(range(87, 97)) + 
	                  list(range(98, 115)) + list(range(116, 126)) + list(range(127, 134)) + 
	                  list(range(135, 143)) + list(range(144, 152)) + list(range(153, 161)) + 
	                  list(range(163, 182)) + list(range(184, 186)))
	
	for row in all_photo_rows:
		remark = remarks.get(str(row), "")
		if remark:
			# DE is merged, so write to D (first column of merge)
			write_to_cell(worksheet, f"D{row}", remark)

		# Helper to fit image to cell size
		def _cell_pixel_size(ws, row_idx, col_idx):
			# Column width: approx pixels = width * 7 (Excel default)
			col_letter = get_column_letter(col_idx)
			col_dim = ws.column_dimensions.get(col_letter)
			col_width = col_dim.width if col_dim and col_dim.width else 8.43
			cell_width_px = int(col_width * 7)

			# Row height: points to pixels (1 pt = 1.333 px)
			row_dim = ws.row_dimensions.get(row_idx)
			row_height = row_dim.height if row_dim and row_dim.height else 15
			cell_height_px = int(row_height * 1.333)
			return cell_width_px, cell_height_px

		row_images = images.get(str(row), [])
		column_index = 6  # Start from F column (column 6)
		max_row_height_px = 0
		for image_path in row_images:
			if not default_storage.exists(image_path):
				continue
			abs_path = default_storage.path(image_path)
			image = ExcelImage(abs_path)
			# Set image size to exactly 2.5" x 2" (240x192 pixels at 96 DPI)
			image.width = 2.5 * 96  # 2.5 inches = 240 pixels
			image.height = 2 * 96    # 2 inches = 192 pixels
			
			img_w = image.width
			img_h = image.height
			
			# Adjust column width to fit image
			if img_w > 0:
				col_letter = get_column_letter(column_index)
				worksheet.column_dimensions[col_letter].width = max(
					worksheet.column_dimensions[col_letter].width or 0,
					img_w / 7,
				)
			if img_h > 0:
				max_row_height_px = max(max_row_height_px, img_h)
			cell = f"{get_column_letter(column_index)}{row}"
			worksheet.add_image(image, cell)
			column_index += 1
		if max_row_height_px > 0:
			worksheet.row_dimensions[row].height = max_row_height_px / 1.333

	# Generate filename with site_id
	site_id_slug = _safe_slug(checklist.site_id, f"site_{checklist.id}")
	filename = f"{site_id_slug}_checklist.xlsx"
	
	copy_name = (
		checklist.template_copy.name
		if checklist.template_copy
		else _build_checklist_path(checklist, filename)
	)
	copy_path = default_storage.path(copy_name)
	os.makedirs(os.path.dirname(copy_path), exist_ok=True)
	workbook.save(copy_path)

	if not checklist.template_copy:
		checklist.template_copy.name = copy_name
		checklist.save(update_fields=["template_copy", "updated_at"])


def _add_template_images(worksheet, template_path: str):
	drawing_path = _get_first_sheet_drawing_path(template_path)
	if not drawing_path:
		return

	with zipfile.ZipFile(template_path) as archive:
		if drawing_path not in archive.namelist():
			return

		rels_path = drawing_path.replace("xl/drawings/", "xl/drawings/_rels/") + ".rels"
		rels_map = {}
		if rels_path in archive.namelist():
			rels_xml = archive.read(rels_path)
			rels_tree = _safe_parse_xml(rels_xml)
			if rels_tree is not None:
				for rel in rels_tree.findall("{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"):
					rels_map[rel.get("Id")] = rel.get("Target")

		drawing_xml = archive.read(drawing_path)
		drawing_tree = _safe_parse_xml(drawing_xml)
		if drawing_tree is None:
			return

		for anchor in drawing_tree.findall("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}twoCellAnchor"):
			from_node = anchor.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}from")
			pic = anchor.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}pic")
			if from_node is None or pic is None:
				continue

			col_node = from_node.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}col")
			row_node = from_node.find("{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}row")
			if col_node is None or row_node is None:
				continue

			col = int(col_node.text or 0) + 1
			row = int(row_node.text or 0) + 1
			cell = f"{get_column_letter(col)}{row}"

			blip = pic.find(
				"{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}blipFill/"
				"{http://schemas.openxmlformats.org/drawingml/2006/main}blip"
			)
			if blip is None:
				continue

			rid = blip.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
			target = rels_map.get(rid)
			if not target:
				continue

			image_path = target.lstrip("/")
			if not image_path.startswith("xl/"):
				image_path = f"xl/{image_path}"
			if image_path not in archive.namelist():
				continue

			width, height = _read_anchor_size(pic)
			image_bytes = archive.read(image_path)
			try:
				pil = PilImage.open(BytesIO(image_bytes))
				image = ExcelImage(pil)
				if width and height:
					image.width = width
					image.height = height
				worksheet.add_image(image, cell)
			except Exception:
				continue


def _get_first_sheet_drawing_path(template_path: str):
	with zipfile.ZipFile(template_path) as archive:
		worksheet_path = "xl/worksheets/sheet1.xml"
		rels_path = "xl/worksheets/_rels/sheet1.xml.rels"
		if worksheet_path not in archive.namelist() or rels_path not in archive.namelist():
			projects = Project.objects.order_by("name")
			return render(
				request,
				"dashboards/team_lead.html",
				{
					"path": path,
					"project": profile.project,
					"checklists": checklists,
					"user_stats": user_stats,
					"project_users": project_users,
					"status_filter": status_filter,
					"user_filter": user_filter,
					"search_query": search_query,
					"locations": locations,
					"work_assignments": work_assignments,
					"engineers": engineers,
					"projects": projects,
				},
			)
	return None


def _read_anchor_size(pic_node):
	try:
		ext = pic_node.find(
			"{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}spPr/"
			"{http://schemas.openxmlformats.org/drawingml/2006/main}xfrm/"
			"{http://schemas.openxmlformats.org/drawingml/2006/main}ext"
		)
		if ext is None:
			return None, None
		cx = int(ext.get("cx", "0"))
		cy = int(ext.get("cy", "0"))
		if not cx or not cy:
			return None, None
		return cx // 9525, cy // 9525
	except Exception:
		return None, None


def _safe_parse_xml(raw: bytes):
	try:
		import xml.etree.ElementTree as ET
		return ET.fromstring(raw)
	except Exception:
		return None


def _safe_slug(value: str, fallback: str):
	clean = slugify(value or "")
	return clean or fallback


def _get_team_lead_name(project: Project):
	team_lead = Profile.objects.filter(project=project, role=Profile.Roles.TEAM_LEAD).first()
	if team_lead:
		return _safe_slug(team_lead.user.username, "team_lead")
	return "unassigned"


def _build_checklist_path(checklist: Checklist, filename: str):
	project_slug = _safe_slug(checklist.project.name, f"project_{checklist.project.id}")
	team_lead_slug = _get_team_lead_name(checklist.project)
	engineer_slug = _safe_slug(checklist.user.username, f"engineer_{checklist.user.id}")
	site_id_slug = _safe_slug(checklist.site_id, f"site_{checklist.id}")
	return (
		f"projects/{project_slug}/team_leads/{team_lead_slug}/"
		f"engineers/{engineer_slug}/sites/{site_id_slug}/{filename}"
	)


def _build_image_path(checklist: Checklist, filename: str):
	project_slug = _safe_slug(checklist.project.name, f"project_{checklist.project.id}")
	team_lead_slug = _get_team_lead_name(checklist.project)
	engineer_slug = _safe_slug(checklist.user.username, f"engineer_{checklist.user.id}")
	site_id_slug = _safe_slug(checklist.site_id, f"site_{checklist.id}")
	return (
		f"projects/{project_slug}/team_leads/{team_lead_slug}/"
		f"engineers/{engineer_slug}/sites/{site_id_slug}/images/{filename}"
	)


@require_http_methods(["POST"])
def location_add(request):
	"""Add a new location pin to the map"""
	if not request.session.get("is_dev_admin") and not (
		request.user.is_authenticated and 
		hasattr(request.user, 'profile') and 
		request.user.profile.role == Profile.Roles.TEAM_LEAD
	):
		messages.error(request, "Permission denied.")
		return redirect("login")
	
	name = request.POST.get("site_id", "").strip() or request.POST.get("name", "").strip()
	latitude = request.POST.get("latitude", "").strip()
	longitude = request.POST.get("longitude", "").strip()
	project_id = request.POST.get("project", "").strip()
	notes = request.POST.get("notes", "").strip()
	
	if not name or not latitude or not longitude:
		messages.error(request, "Site ID, latitude and longitude are required.")
	else:
		try:
			lat = float(latitude)
			lon = float(longitude)
			
			location = GeoLocation.objects.create(
				name=name,
				latitude=lat,
				longitude=lon,
				project_id=project_id if project_id else None,
				created_by=request.user if request.user.is_authenticated else User.objects.first(),
				notes=notes
			)
			messages.success(request, f"Location '{name}' added successfully!")
		except ValueError:
			messages.error(request, "Invalid latitude or longitude format.")
		except Exception as e:
			messages.error(request, f"Error adding location: {str(e)}")
	
	if request.session.get("is_dev_admin"):
		return redirect("dev_admin")
	return redirect("user_dashboard", path=request.user.profile.path)


@require_http_methods(["POST"])
def location_import(request):
	"""Import location pins from an Excel file (columns: Site ID, Latitude, Longitude)."""
	if not request.session.get("is_dev_admin") and not (
		request.user.is_authenticated and 
		hasattr(request.user, 'profile') and 
		request.user.profile.role == Profile.Roles.TEAM_LEAD
	):
		messages.error(request, "Permission denied.")
		return redirect("login")

	upload = request.FILES.get("locations_file")
	if not upload:
		messages.error(request, "Please select an Excel file.")
		if request.session.get("is_dev_admin"):
			return redirect("dev_admin")
		return redirect("user_dashboard", path=request.user.profile.path)

	try:
		from openpyxl import load_workbook
		wb = load_workbook(upload, data_only=True)
		ws = wb.active

		created = 0
		for row in ws.iter_rows(min_row=1, values_only=True):
			if not row or len(row) < 3:
				continue
			site_id, lat_val, lon_val = row[0], row[1], row[2]
			if site_id is None or lat_val is None or lon_val is None:
				continue
			try:
				lat = float(lat_val)
				lon = float(lon_val)
			except (TypeError, ValueError):
				# Skip header or invalid rows
				continue

			GeoLocation.objects.create(
				name=str(site_id).strip(),
				latitude=lat,
				longitude=lon,
				project=None,
				created_by=request.user if request.user.is_authenticated else User.objects.first(),
				notes="",
			)
			created += 1

		messages.success(request, f"Imported {created} locations.")
	except Exception as e:
		messages.error(request, f"Import failed: {str(e)}")

	if request.session.get("is_dev_admin"):
		return redirect("dev_admin")
	return redirect("user_dashboard", path=request.user.profile.path)


@require_http_methods(["POST"])
def location_delete(request, location_id: int):
	"""Delete a location pin from the map (Admin only)"""
	if not request.session.get("is_dev_admin"):
		messages.error(request, "Permission denied. Only admins can delete locations.")
		return redirect("login")
	
	location = get_object_or_404(GeoLocation, id=location_id)
	location_name = location.name
	location.delete()
	messages.success(request, f"Location '{location_name}' deleted successfully!")
	
	return redirect("dev_admin")


@require_http_methods(["POST"])
def location_delete_all(request):
	"""Delete all location pins from the map (Admin only)"""
	if not request.session.get("is_dev_admin"):
		messages.error(request, "Permission denied. Only admins can delete locations.")
		return redirect("login")

	GeoLocation.objects.all().delete()
	messages.success(request, "All locations deleted.")

	return redirect("dev_admin")


@require_http_methods(["POST"])
def assign_work(request):
	"""Create a new work assignment for an engineer"""
	from django.utils import timezone
	
	# Check permissions
	if not request.session.get("is_dev_admin") and not (
		request.user.is_authenticated and 
		hasattr(request.user, 'profile') and 
		request.user.profile.role == Profile.Roles.TEAM_LEAD
	):
		messages.error(request, "Permission denied.")
		return redirect("login")
	
	try:
		site_id = request.POST.get("site_id", "").strip()
		latitude = request.POST.get("latitude", "").strip()
		longitude = request.POST.get("longitude", "").strip()
		description = request.POST.get("description", "").strip()
		engineer_id = request.POST.get("engineer_id")
		project_id = request.POST.get("project_id")
		
		# Validation (description is optional)
		if not all([site_id, latitude, longitude, engineer_id, project_id]):
			messages.error(request, "Site ID, latitude, longitude, engineer and project are required.")
			if request.session.get("is_dev_admin"):
				return redirect("dev_admin")
			return redirect("user_dashboard", path=request.user.profile.path)
		
		# Validate coordinates
		try:
			lat = float(latitude)
			lng = float(longitude)
			if not (-90 <= lat <= 90) or not (-180 <= lng <= 180):
				raise ValueError("Invalid coordinate range")
		except ValueError:
			messages.error(request, "Invalid coordinates. Latitude must be between -90 and 90, longitude between -180 and 180.")
			if request.session.get("is_dev_admin"):
				return redirect("dev_admin")
			return redirect("user_dashboard", path=request.user.profile.path)
		
		# Get engineer and project
		engineer = get_object_or_404(User, id=engineer_id)
		project = get_object_or_404(Project, id=project_id)
		
		# Check if this site is already assigned to anyone
		existing_assignment = WorkAssignment.objects.filter(site_id=site_id).first()
		if existing_assignment:
			messages.error(
				request, 
				f"Site '{site_id}' is already assigned to {existing_assignment.assigned_to.username}. "
				f"Each site can only be assigned once."
			)
			if request.session.get("is_dev_admin"):
				return redirect("dev_admin")
			return redirect("user_dashboard", path=request.user.profile.path)
		
		# Get the user who is assigning
		if request.session.get("is_dev_admin"):
			assigned_by = User.objects.filter(is_superuser=True).first()
		else:
			assigned_by = request.user
		
		# Create work assignment (checklist will be created by engineer)
		try:
			work = WorkAssignment.objects.create(
				site_id=site_id,
				latitude=lat,
				longitude=lng,
				description=description,
				assigned_to=engineer,
				assigned_by=assigned_by,
				project=project,
				status=WorkAssignment.Status.PENDING
			)

			# Create checklist immediately so team lead/admin can see it
			answer_data = {"12": site_id}
			checklist = Checklist.objects.create(
				user=engineer,
				project=project,
				site_id=site_id,
				status=Checklist.Status.DRAFT,
				answer_data=answer_data,
			)
			work.checklist = checklist
			work.save(update_fields=["checklist"])
			if project.template_file:
				_create_or_update_excel_copy(checklist)
			
			messages.success(request, f"Work assigned to {engineer.username} successfully!")
		except IntegrityError:
			messages.error(
				request, 
				f"Site '{site_id}' is already assigned. Each site can only be assigned once."
			)
			if request.session.get("is_dev_admin"):
				return redirect("dev_admin")
			return redirect("user_dashboard", path=request.user.profile.path)
		
	except Exception as e:
		messages.error(request, f"Error assigning work: {str(e)}")
	
	if request.session.get("is_dev_admin"):
		return redirect("dev_admin")
	return redirect("user_dashboard", path=request.user.profile.path)


@require_http_methods(["POST"])
def work_edit(request, work_id: int):
	"""Edit an assigned work (team lead can edit their own assignments, admin can edit all)."""
	if not request.session.get("is_dev_admin") and not (
		request.user.is_authenticated and 
		hasattr(request.user, 'profile') and 
		request.user.profile.role == Profile.Roles.TEAM_LEAD
	):
		messages.error(request, "Permission denied.")
		return redirect("login")

	work = get_object_or_404(WorkAssignment, id=work_id)

	if not request.session.get("is_dev_admin"):
		# Team leader can only edit work they assigned
		if work.assigned_by != request.user:
			messages.error(request, "You can only edit work you assigned.")
			return redirect("user_dashboard", path=request.user.profile.path)

	site_id = request.POST.get("site_id", "").strip()
	latitude = request.POST.get("latitude", "").strip()
	longitude = request.POST.get("longitude", "").strip()
	description = request.POST.get("description", "").strip()
	engineer_id = request.POST.get("engineer_id")

	if not all([site_id, latitude, longitude, engineer_id]):
		messages.error(request, "Site ID, latitude, longitude and engineer are required.")
		if request.session.get("is_dev_admin"):
			return redirect("dev_admin")
		return redirect("user_dashboard", path=request.user.profile.path)

	try:
		lat = float(latitude)
		lon = float(longitude)
	except ValueError:
		messages.error(request, "Invalid latitude or longitude format.")
		if request.session.get("is_dev_admin"):
			return redirect("dev_admin")
		return redirect("user_dashboard", path=request.user.profile.path)

	engineer = get_object_or_404(User, id=engineer_id)
	work.site_id = site_id
	work.latitude = lat
	work.longitude = lon
	work.description = description
	work.assigned_to = engineer
	work.save()

	# Keep checklist site_id in sync if it exists
	if work.checklist:
		answers = work.checklist.answer_data or {}
		answers["12"] = site_id
		work.checklist.site_id = site_id
		work.checklist.answer_data = answers
		work.checklist.save(update_fields=["site_id", "answer_data", "updated_at"])
		_create_or_update_excel_copy(work.checklist)

	messages.success(request, "Work updated.")
	if request.session.get("is_dev_admin"):
		return redirect("dev_admin")
	return redirect("user_dashboard", path=request.user.profile.path)


@require_http_methods(["POST"])
def work_delete(request, work_id: int):
	"""Delete an assigned work (admin only)."""
	if not request.session.get("is_dev_admin"):
		messages.error(request, "Permission denied.")
		return redirect("login")

	work = get_object_or_404(WorkAssignment, id=work_id)
	work.delete()
	messages.success(request, "Work deleted.")
	return redirect("dev_admin")


@require_http_methods(["POST"])
def update_work_status(request, work_id: int):
	"""Update work assignment status (for engineers)"""
	from django.utils import timezone
	
	if not request.user.is_authenticated:
		messages.error(request, "Please log in.")
		return redirect("login")
	
	work = get_object_or_404(WorkAssignment, id=work_id, assigned_to=request.user)
	new_status = request.POST.get("status")
	engineer_notes = request.POST.get("engineer_notes", "").strip()
	
	if new_status in [WorkAssignment.Status.IN_PROGRESS, WorkAssignment.Status.SUBMITTED]:
		work.status = new_status
		if engineer_notes:
			work.engineer_notes = engineer_notes
		
		if new_status == WorkAssignment.Status.IN_PROGRESS and not work.started_at:
			work.started_at = timezone.now()
		elif new_status == WorkAssignment.Status.SUBMITTED and not work.submitted_at:
			work.submitted_at = timezone.now()
		
		work.save()
		messages.success(request, f"Work status updated to {work.get_status_display()}!")
	else:
		messages.error(request, "Invalid status.")
	
	return redirect("user_dashboard", path=request.user.profile.path)


@require_http_methods(["POST"])
def complete_work(request, work_id: int):
	"""Mark work as completed (for admin/team lead)"""
	from django.utils import timezone
	
	# Check permissions
	if not request.session.get("is_dev_admin") and not (
		request.user.is_authenticated and 
		hasattr(request.user, 'profile') and 
		request.user.profile.role == Profile.Roles.TEAM_LEAD
	):
		messages.error(request, "Permission denied.")
		return redirect("login")
	
	work = get_object_or_404(WorkAssignment, id=work_id)
	work.status = WorkAssignment.Status.COMPLETED
	work.completed_at = timezone.now()
	work.save()
	
	messages.success(request, f"Work for {work.site_id} marked as completed!")
	
	if request.session.get("is_dev_admin"):
		return redirect("dev_admin")
	return redirect("user_dashboard", path=request.user.profile.path)


@require_http_methods(["GET"])
def create_checklist_from_work(request, work_id: int):
	"""Create a checklist from work assignment and pre-fill site_id in row 12"""
	if not request.user.is_authenticated:
		messages.error(request, "Please log in.")
		return redirect("login")
	
	work = get_object_or_404(WorkAssignment, id=work_id, assigned_to=request.user)
	
	# Check if checklist already exists
	if work.checklist:
		messages.info(request, "Checklist already exists for this work.")
		return redirect("engineer_checklist_edit", path=request.user.profile.path, checklist_id=work.checklist.id)
	
	# Create checklist with site_id pre-filled in row 12
	answer_data = {}
	answer_data["12"] = work.site_id  # Pre-fill row 12 with site_id
	
	checklist = Checklist.objects.create(
		user=request.user,
		project=work.project,
		site_id=work.site_id,
		status=Checklist.Status.DRAFT,
		answer_data=answer_data,
		comment=f"Created from work assignment: {work.description[:100]}"
	)
	
	# Create Excel copy from project template
	_create_or_update_excel_copy(checklist)
	
	# Link checklist to work assignment
	work.checklist = checklist
	work.save()
	
	messages.success(request, f"Checklist created with Site ID pre-filled!")
	return redirect("engineer_checklist_edit", path=request.user.profile.path, checklist_id=checklist.id)

# ===== NEW DETAILED CHECKLIST VIEWS =====

@require_http_methods(["GET"])
def checklist_detail_view(request, checklist_id):
	"""Main view for detailed checklist form"""
	checklist = get_object_or_404(Checklist, id=checklist_id)
	
	# Check permissions
	profile = getattr(request.user, "profile", None)
	is_team_lead = profile and profile.role == Profile.Roles.TEAM_LEAD
	is_same_project = is_team_lead and checklist.project == profile.project
	is_admin = request.session.get("is_dev_admin")
	is_engineer = request.user == checklist.user
	
	if not (is_engineer or is_admin or is_same_project):
		raise Http404("Not authorized")
	
	# Determine if user can edit (engineers can't edit FINAL checklists, but team leads and admins can)
	can_edit = is_admin or is_team_lead or (is_engineer and checklist.status != Checklist.Status.FINAL)
	
	# Get template path
	if checklist.template_copy:
		template_path = checklist.template_copy.path
	elif checklist.project.template_file:
		template_path = checklist.project.template_file.path
	else:
		messages.error(request, "No template file available for this checklist.")
		return redirect("user_dashboard", path=request.user.profile.path)
	
	# Read template questions
	general_questions, image_questions, dc_power_questions, _site_id = _read_template_questions(template_path)
	
	# Organize questions by section
	section_questions = {
		'general': general_questions,
		'civil': [q for q in image_questions if 22 <= q['row'] <= 70],
		'electromechanical': [q for q in image_questions if 73 <= q['row'] <= 84],
		'powersupply': [q for q in image_questions if 87 <= q['row'] <= 96],
		'dgsg': [q for q in image_questions if 98 <= q['row'] <= 114],
		'hybrid': [q for q in image_questions if 116 <= q['row'] <= 125],
		'solar': [q for q in image_questions if 127 <= q['row'] <= 133],
		'transformer': [q for q in image_questions if 135 <= q['row'] <= 142],
		'service': [q for q in image_questions if 144 <= q['row'] <= 151],
		'shareable': [q for q in image_questions if 153 <= q['row'] <= 160],
		'shelter': [q for q in image_questions if 163 <= q['row'] <= 181],
		'extra': [q for q in image_questions if 184 <= q['row'] <= 185],
		'dcpower': dc_power_questions,
	}
	
	# Get existing answers
	answers = checklist.answer_data or {}
	remarks = checklist.remark_data or {}
	images = checklist.image_data or {}
	
	print(f"\n{'='*60}")
	print(f"LOADING CHECKLIST {checklist_id}")
	print(f"Total images in database: {len(images)} rows with images")
	for row, img_list in images.items():
		print(f"  Row {row}: {len(img_list)} images")
	print(f"{'='*60}\n")
	
	# Add existing values to questions
	for section_key, questions_list in section_questions.items():
		for q in questions_list:
			q['value'] = answers.get(str(q['row']), '')
			q['remark'] = remarks.get(str(q['row']), '')
			# Add MEDIA_URL to image paths
			from django.conf import settings
			image_paths = images.get(str(q['row']), [])
			q['images'] = [settings.MEDIA_URL + path if not path.startswith('/media') else path for path in image_paths]
	
	# Section list for template loop
	section_list = [
		('powersupply', 'Power Supply'),
		('dgsg', 'DG/SG Set'),
		('hybrid', 'Hybrid AC/DC'),
		('solar', 'Solar'),
		('transformer', 'Transformer'),
		('service', 'Service Disconnect'),
		('shareable', 'SHAREABLE MDB/LDB'),
		('shelter', 'SHELTER/ODU'),
		('extra', 'Extra'),
		('dcpower', 'DC Power System'),
	]
	
	return render(request, "dashboards/checklist_detail.html", {
		"checklist": checklist,
		"section_questions": section_questions,
		"section_list": section_list,
		"can_edit": can_edit,
	})


@require_http_methods(["GET"])
def checklist_data_api(request, checklist_id):
	"""API endpoint to get all checklist data"""
	from .models import ChecklistSection, ChecklistImage, DCPowerSystemData, TowerEquipment, ElectricalData
	
	checklist = get_object_or_404(Checklist, id=checklist_id)
	
	# Get all sections
	sections = {}
	for section in checklist.sections.all():
		if section.section_name not in sections:
			sections[section.section_name] = []
		
		images = list(section.images.values('id', 'image', 'column_position'))
		
		sections[section.section_name].append({
			'id': section.id,
			'row_number': section.row_number,
			'question': section.question,
			'answer': section.answer,
			'remarks': section.remarks,
			'images': images
		})
	
	# Get DC Power data from answer_data
	answers = checklist.answer_data or {}
	dc_power = []
	
	# Get Tower Equipment from answer_data
	tower_equipment = {}
	electrical = {}
	
	for key, value in answers.items():
		if key.startswith('equipment_'):
			# Parse equipment data
			operator = value.get('operator', '')
			equip_type = value.get('type', '')
			equip_data = value.get('data', {})
			
			equip_key = f"{operator}_{equip_type}"
			if equip_key not in tower_equipment:
				tower_equipment[equip_key] = []
			tower_equipment[equip_key].append(equip_data)
			
		elif key.startswith('electrical_'):
			# Parse electrical data
			row_num = key.replace('electrical_', '')
			electrical[row_num] = value

	zip_upload = answers.get('zip_upload')
	
	return JsonResponse({
		'sections': sections,
		'dc_power': dc_power,
		'tower_equipment': tower_equipment,
		'electrical': electrical,
		'zip_upload': zip_upload
	})


@require_http_methods(["POST"])
def checklist_autosave_api(request, checklist_id):
	"""Auto-save endpoint for checklist data"""
	import json
	
	checklist = get_object_or_404(Checklist, id=checklist_id)
	
	try:
		data = json.loads(request.body)
		
		# Handle answer save
		if 'answer' in data and 'row' in data:
			row = str(data['row'])
			answer_data = checklist.answer_data or {}
			answer_data[row] = data['answer']
			checklist.answer_data = answer_data
			checklist.save()
			return JsonResponse({'status': 'success'})
		
		# Handle remark save
		if 'remark' in data and 'row' in data:
			row = str(data['row'])
			remark_data = checklist.remark_data or {}
			remark_data[row] = data['remark']
			checklist.remark_data = remark_data
			checklist.save()
			return JsonResponse({'status': 'success'})
		
		# Handle tower equipment save
		if data.get('save_type') == 'tower_equipment':
			operator = data.get('operator')
			equipment_type = data.get('equipment_type')
			unique_id = data.get('unique_id')
			equipment_data = data.get('equipment_data', {})
			
			# Store in answer_data JSON for now (simple approach)
			answer_data = checklist.answer_data or {}
			equipment_key = f"equipment_{unique_id}"
			answer_data[equipment_key] = {
				'operator': operator,
				'type': equipment_type,
				'data': equipment_data
			}
			checklist.answer_data = answer_data
			checklist.save()
			return JsonResponse({'status': 'success'})
		
		# Handle electrical data save
		if data.get('save_type') == 'electrical_data':
			row_number = data.get('row')
			electrical_data = data.get('electrical_data', {})
			
			# Store in answer_data JSON
			answer_data = checklist.answer_data or {}
			answer_data[f"electrical_{row_number}"] = electrical_data
			checklist.answer_data = answer_data
			checklist.save()
			return JsonResponse({'status': 'success'})
		
		# Old format handling for backward compatibility
		save_type = data.get('type')
		
		if save_type == 'section':
			section_data = data.get('data')
			section, created = ChecklistSection.objects.update_or_create(
				checklist=checklist,
				section_name=section_data['section_name'],
				row_number=section_data['row_number'],
				defaults={
					'question': section_data.get('question', ''),
					'answer': section_data.get('answer', ''),
					'remarks': section_data.get('remarks', '')
				}
			)
			return JsonResponse({'status': 'success', 'section_id': section.id})
		
		elif save_type == 'dc_power':
			dc_data = data.get('data')
			dc_obj, created = DCPowerSystemData.objects.update_or_create(
				checklist=checklist,
				row_number=dc_data['row_number'],
				defaults={
					'field_label': dc_data['field_label'],
					'field_value': dc_data.get('field_value', '')
				}
			)
			return JsonResponse({'status': 'success', 'id': dc_obj.id})
		
		return JsonResponse({'status': 'error', 'message': 'Invalid save type'}, status=400)
	
	except Exception as e:
		return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@require_http_methods(["POST"])
def checklist_upload_image(request, checklist_id):
	"""Handle image upload for checklist sections"""
	from PIL import Image as PilImage
	import io
	import os
	from django.core.files.base import ContentFile
	
	print(f"\n{'='*60}")
	print(f"UPLOAD IMAGE REQUEST - Checklist ID: {checklist_id}")
	print(f"Method: {request.method}")
	print(f"User: {request.user}")
	print(f"POST data: {dict(request.POST)}")
	print(f"FILES: {list(request.FILES.keys())}")
	print(f"{'='*60}\n")
	
	# Check authentication and permissions
	checklist = get_object_or_404(Checklist, id=checklist_id)
	
	# Allow dev admin, checklist owner, or team lead from same project
	profile = getattr(request.user, "profile", None)
	is_admin = request.session.get("is_dev_admin")
	is_owner = request.user == checklist.user
	is_team_lead = profile and profile.role == Profile.Roles.TEAM_LEAD and checklist.project == profile.project
	
	print(f"Permissions - Admin: {is_admin}, Owner: {is_owner}, Team Lead: {is_team_lead}")
	
	if not (is_admin or is_owner or is_team_lead):
		print("❌ Permission denied")
		return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)
	
	# Check if user can edit (not FINAL status for engineers)
	if is_owner and not is_admin and not is_team_lead:
		if checklist.status == Checklist.Status.FINAL:
			print("❌ Cannot edit FINAL checklist")
			return JsonResponse({'status': 'error', 'message': 'Cannot edit finalized checklist'}, status=403)
	
	row = request.POST.get('row')
	
	if not row:
		print("❌ No row number provided")
		return JsonResponse({'status': 'error', 'message': 'Row number required'}, status=400)
	
	print(f"Row number: {row}")
	
	uploaded_files = request.FILES.getlist('images')
	print(f"Uploaded files count: {len(uploaded_files)}")
	
	if not uploaded_files:
		print("❌ No images provided")
		return JsonResponse({'status': 'error', 'message': 'No images provided'}, status=400)
	
	try:
		image_data = checklist.image_data or {}
		row_images = image_data.get(str(row), [])
		new_images = []  # Track newly added images
		
		print(f"Processing {len(uploaded_files)} image(s)...")
		
		for uploaded_file in uploaded_files:
			print(f"Processing file: {uploaded_file.name}, size: {uploaded_file.size} bytes")
			
			# Save image at full resolution without resizing
			# Generate unique filename with timestamp
			import time
			ext = os.path.splitext(uploaded_file.name)[1] or '.jpg'
			timestamp = int(time.time() * 1000)
			filename = f"checklist_{checklist.id}_row_{row}_{timestamp}{ext}"
			
			# Save file using the custom path builder
			from django.core.files.storage import default_storage
			file_path = _build_image_path(checklist, filename)
			
			# Reset file pointer to beginning
			uploaded_file.seek(0)
			file_path = default_storage.save(file_path, uploaded_file)
			
			print(f"✅ Saved image to: {file_path}")
			
			# Add to image list
			row_images.append(file_path)
			new_images.append(file_path)
		
		image_data[str(row)] = row_images
		checklist.image_data = image_data
		checklist.save()
		
		print(f"✅ SUCCESS - Saved {len(new_images)} new image(s)")
		print(f"Response: {{'status': 'success', 'images': {len(row_images)} total, 'new_images': {new_images}}}")
		
		return JsonResponse({'status': 'success', 'images': row_images, 'new_images': new_images})
	
	except Exception as e:
		import traceback
		traceback.print_exc()
		return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@require_http_methods(["POST"])
def checklist_delete_image(request):
	"""Delete a checklist image"""
	import json
	
	print(f"\n{'='*60}")
	print(f"DELETE IMAGE REQUEST")
	print(f"User: {request.user}")
	
	try:
		data = json.loads(request.body)
		checklist_id = data.get('checklist_id')
		image_path = data.get('image_path')
		row = str(data.get('row'))
		
		print(f"Checklist ID: {checklist_id}")
		print(f"Image path: {image_path}")
		print(f"Row: {row}")
		
		# Clean the image path - remove /media/ prefix if present
		if image_path.startswith('/media/'):
			image_path = image_path.replace('/media/', '', 1)
		
		print(f"Cleaned path: {image_path}")
		
		checklist = get_object_or_404(Checklist, id=checklist_id)
		
		# Check permissions
		profile = getattr(request.user, "profile", None)
		is_admin = request.session.get("is_dev_admin")
		is_owner = request.user == checklist.user
		is_team_lead = profile and profile.role == Profile.Roles.TEAM_LEAD and checklist.project == profile.project
		
		if not (is_admin or is_owner or is_team_lead):
			print("❌ Permission denied")
			return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)
		
		image_data = checklist.image_data or {}
		
		print(f"Current images in row {row}: {image_data.get(row, [])}")
		
		if row in image_data and image_path in image_data[row]:
			image_data[row].remove(image_path)
			checklist.image_data = image_data
			checklist.save()
			
			print(f"✅ Removed from database")
			
			# Delete physical file
			from django.core.files.storage import default_storage
			if default_storage.exists(image_path):
				default_storage.delete(image_path)
				print(f"✅ Deleted physical file")
			else:
				print(f"⚠️ Physical file not found: {image_path}")
		else:
			print(f"❌ Image not found in database")
		
		print(f"{'='*60}\n")
		return JsonResponse({'status': 'success'})
	except Exception as e:
		return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@require_http_methods(["POST"])
def checklist_upload_zip(request, checklist_id):
	"""Upload a ZIP file for the checklist."""
	from django.core.files.storage import default_storage
	from django.utils.text import slugify

	checklist = get_object_or_404(Checklist, id=checklist_id)
	zip_file = request.FILES.get('zip_file')
	if not zip_file:
		return JsonResponse({'status': 'error', 'message': 'No file provided'}, status=400)

	answers = checklist.answer_data or {}
	old_zip = answers.get('zip_upload', {})
	old_path = old_zip.get('path') if isinstance(old_zip, dict) else None
	if old_path and default_storage.exists(old_path):
		default_storage.delete(old_path)

	safe_name = slugify(zip_file.name.rsplit('.', 1)[0]) or 'checklist'
	ext = zip_file.name.split('.')[-1]
	filename = f"{safe_name}.{ext}"
	file_path = f"checklist_zips/{checklist.id}/{filename}"
	saved_path = default_storage.save(file_path, zip_file)

	answers['zip_upload'] = {
		'path': saved_path,
		'name': zip_file.name,
		'size': zip_file.size,
	}
	checklist.answer_data = answers
	checklist.save(update_fields=["answer_data", "updated_at"])

	return JsonResponse({'status': 'success', 'zip': answers['zip_upload']})


@require_http_methods(["GET"])
def checklist_download_zip(request, checklist_id):
	"""Download uploaded ZIP file for checklist."""
	from django.core.files.storage import default_storage

	checklist = get_object_or_404(Checklist, id=checklist_id)
	zip_info = (checklist.answer_data or {}).get('zip_upload') or {}
	zip_path = zip_info.get('path')
	if not zip_path or not default_storage.exists(zip_path):
		raise Http404("ZIP file not found.")

	file_handle = default_storage.open(zip_path, "rb")
	filename = zip_info.get('name') or os.path.basename(zip_path)
	return FileResponse(file_handle, as_attachment=True, filename=filename)


@require_http_methods(["POST"])
def checklist_delete_equipment(request, equipment_id):
	"""Delete tower equipment row"""
	from .models import TowerEquipment
	
	try:
		equipment = get_object_or_404(TowerEquipment, id=equipment_id)
		equipment.delete()
		return JsonResponse({'status': 'success'})
	except Exception as e:
		return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@require_http_methods(["POST"])
def checklist_delete_electrical(request, electrical_id):
	"""Delete electrical data row"""
	from .models import ElectricalData
	
	try:
		electrical = get_object_or_404(ElectricalData, id=electrical_id)
		electrical.delete()
		return JsonResponse({'status': 'success'})
	except Exception as e:
		return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@require_http_methods(["GET", "POST"])
def checklist_submit(request, checklist_id):
	"""Submit checklist and update Excel with all data"""
	from django.utils import timezone
	
	checklist = get_object_or_404(Checklist, id=checklist_id)
	
	# Check permissions
	if request.user != checklist.user and not request.session.get("is_dev_admin"):
		messages.error(request, "Not authorized")
		return redirect("login")
	
	# Update Excel with all current data
	_create_or_update_excel_copy(checklist)
	
	# Update status to SUBMITTED
	checklist.status = Checklist.Status.SUBMITTED
	checklist.save(update_fields=["status", "updated_at"])
	
	# Update linked work assignment if exists
	try:
		work_assignment = WorkAssignment.objects.get(checklist=checklist)
		work_assignment.status = WorkAssignment.Status.SUBMITTED
		work_assignment.submitted_at = timezone.now()
		work_assignment.save()
	except WorkAssignment.DoesNotExist:
		pass
	
	messages.success(request, "Checklist submitted successfully!")
	
	# Redirect back to user dashboard
	if hasattr(request.user, 'profile') and request.user.profile.path:
		return redirect("user_dashboard", path=request.user.profile.path)
	return redirect("dev_admin")
